from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class ScreenerExporter:

    @staticmethod
    def colour_file(filename):

        wb = load_workbook(filename)

        green = PatternFill(
            fill_type="solid",
            fgColor="90EE90"
        )

        red = PatternFill(
            fill_type="solid",
            fgColor="FFC7CE"
        )

        for ws in wb.worksheets:

            for row in ws.iter_rows(min_row=2):

                for cell in row:

                    if isinstance(cell.value, (int, float)):

                        if cell.value > 0:
                            cell.fill = green

                        else:
                            cell.fill = red

        wb.save(filename)