from openpyxl import load_workbook
from openpyxl.styles import PatternFill

FILE = "output/peer_comparison.xlsx"

wb = load_workbook(FILE)

green = PatternFill(
    fill_type="solid",
    fgColor="90EE90"
)

yellow = PatternFill(
    fill_type="solid",
    fgColor="FFF59D"
)

red = PatternFill(
    fill_type="solid",
    fgColor="FFCDD2"
)

gold = PatternFill(
    fill_type="solid",
    fgColor="FFD54F"
)

for ws in wb.worksheets:

    headers = [c.value for c in ws[1]]

    percentile_cols = [
        i + 1
        for i, h in enumerate(headers)
        if str(h).endswith("_pct")
    ]

    company_col = headers.index("company_id") + 1

    for row in range(2, ws.max_row + 1):

        company = ws.cell(row, company_col).value

        # Highlight first company as benchmark
        if row == 2:
            for cell in ws[row]:
                cell.fill = gold

        for col in percentile_cols:

            cell = ws.cell(row, col)

            if not isinstance(cell.value, (int, float)):
                continue

            if cell.value >= 0.75:
                cell.fill = green
            elif cell.value >= 0.25:
                cell.fill = yellow
            else:
                cell.fill = red

wb.save(FILE)

print("Peer report formatted.")